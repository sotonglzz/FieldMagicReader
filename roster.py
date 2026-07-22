"""Excel roster loader: used only to FILTER time-based staff matches.

The human-maintained schedule spreadsheet is half-filled and unreliable, so it
never creates matches on its own. When a job has rostered staff for an
install/removal, only those people (resolved via nickname aliases) are kept
from the time-based candidates. When the roster has no staff for a job, the
time matcher is left unchanged.
"""

from __future__ import annotations

import json
import logging
import os
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime
from glob import glob
from threading import RLock

logger = logging.getLogger(__name__)

DB_NAME = "jobs_cache.db"
ALIASES_PATH = "roster_aliases.json"
# Prefer an exact name; fall back to the newest Schedule*.xlsx in the project root.
DEFAULT_ROSTER_FILENAME = "Schedule July 2025 - June 2026 - copy.xlsx"

MONTH_SHEETS = (
    "July", "August", "September", "October", "November", "December",
    "January", "February", "March", "April", "May", "June",
)

# July–December belong to the starting year; January–June to starting_year+1.
_MONTH_YEAR_OFFSET = {
    "July": 0, "August": 0, "September": 0, "October": 0,
    "November": 0, "December": 0,
    "January": 1, "February": 1, "March": 1, "April": 1,
    "May": 1, "June": 1,
}

_MONTH_NUM = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

JOBNUM_RE = re.compile(r"\(\s*(Q?\d{4,6}(?:-\d+)?(?:,\d+)*)\s*\)", re.IGNORECASE)
DAY_HEADER_RE = re.compile(
    r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),\s+"
    r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)",
    re.IGNORECASE,
)
VEHICLE_HINTS = re.compile(
    r"^(isuzu|fuso|ym|yn|box\d*|fb\d*|autovan|newvan|minivan|van|structure|"
    r"marq|wt|stillages?|1box|1van|2xbox|8fb)$",
    re.I,
)
DATE_LIKE = re.compile(
    r"^(mon|tue|tues|wed|weds|thu|thur|thurs|fri|sat|sun)\b|\d{1,2}/\d{1,2}",
    re.I,
)
JUNK_TOKENS = {
    "at", "back", "till", "until", "drivers", "driver", "staff", "minimum",
    "after", "before", "am", "pm", "w", "x", "and", "with", "the", "to",
    "from", "for", "leave", "park", "outside", "going", "area", "done",
    "by", "change", "should", "shoud", "tba", "tbc", "na", "n/a",
}
SKIP_SUBURBS = {
    "warehouse", "braeside", "office", "cfh", "tennis",
    "warehouse + braeside", "matho", "jag",
}

# Excel Type column -> install/removal. Unknown/blank kinds match either.
_KIND_MAP = {
    "setup": "install",
    "onsite": "install",
    "site": "install",
    "site visit": "install",
    "deliver": "install",
    "dropoff": "install",
    "pulldown": "removal",
    "pulldowm": "removal",
    "remove": "removal",
    "return": "removal",
}

# Pickup/collect rows are DIY and never filter on-site installs.
_IGNORE_TYPES = {"pickup", "pickuo", "collect", "interview", "meeting", "zoom", "walkthru"}


class RosterIndex:
    """In-memory index of rostered staff per job/kind/date."""

    def __init__(self):
        # normalised job number -> list of entries
        # entry: {"date": date, "kind": "install"|"removal"|None, "staff": [tokens], "suburb": str}
        self.by_job = defaultdict(list)
        self.nicknames = set()  # raw Excel staff tokens
        self.source_path = None
        self.loaded_at = None
        self.entry_count = 0
        self.load_error = None  # human-readable reason when load produced nothing

    def add(self, job_number, day, kind, staff_tokens, suburb=""):
        if not job_number or not staff_tokens:
            return
        for key in _job_keys(job_number):
            self.by_job[key].append(
                {
                    "date": day,
                    "kind": kind,
                    "staff": list(staff_tokens),
                    "suburb": suburb,
                }
            )
        self.nicknames.update(staff_tokens)
        self.entry_count += 1

    def rostered_staff(self, job_text, kind, anchor_date, day_slop=1):
        """Return set of Excel nickname tokens rostered for this job event.

        Empty set means "no roster signal" — caller must not filter.
        """
        if not job_text or not anchor_date:
            return set()
        keys = _job_keys(job_text)
        found = set()
        for key in keys:
            for entry in self.by_job.get(key, ()):
                if entry["kind"] and kind and entry["kind"] != kind:
                    continue
                if entry["date"] is None:
                    continue
                if abs((entry["date"] - anchor_date).days) > day_slop:
                    continue
                found.update(entry["staff"])
        return found


def _job_keys(value):
    """Generate normalised job-number keys for join lookups."""
    raw = str(value or "").strip().upper()
    if not raw:
        return set()
    keys = {raw}
    core = re.sub(r"^Q", "", raw)
    main = core.split(",")[0].split("-")[0]
    if main.isdigit():
        keys.add(main)
        keys.add(main.zfill(6))
        keys.add(main.lstrip("0") or "0")
    return keys


def find_roster_path(explicit=None):
    if explicit and os.path.isfile(explicit):
        return explicit
    if os.path.isfile(DEFAULT_ROSTER_FILENAME):
        return DEFAULT_ROSTER_FILENAME
    candidates = sorted(glob("Schedule*.xlsx"), key=os.path.getmtime, reverse=True)
    return candidates[0] if candidates else None


def load_aliases(path=ALIASES_PATH):
    if not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("Could not read roster aliases from %s: %s", path, exc)
        return {}
    if not isinstance(data, dict):
        return {}
    # nickname (original case preserved as key used in Excel) -> full timesheet name
    return {
        str(k).strip(): str(v).strip()
        for k, v in data.items()
        if str(k).strip() and str(v).strip()
    }


def save_aliases(aliases, path=ALIASES_PATH):
    cleaned = {
        str(k).strip(): str(v).strip()
        for k, v in (aliases or {}).items()
        if str(k).strip() and str(v).strip()
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(cleaned, handle, indent=2, ensure_ascii=False, sort_keys=True)
    return cleaned


def list_timesheet_employees(db_name=DB_NAME):
    """Sorted unique team_member values from the timesheets table."""
    try:
        with sqlite3.connect(db_name) as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT team_member
                FROM timesheets
                WHERE team_member IS NOT NULL AND TRIM(team_member) != ''
                ORDER BY team_member COLLATE NOCASE
                """
            ).fetchall()
    except sqlite3.OperationalError:
        return []
    return [row[0] for row in rows]


def _parse_day_header(text, sheet_name, start_year):
    match = DAY_HEADER_RE.match(str(text).strip())
    if not match:
        return None
    day_num = int(match.group(2))
    month_name = match.group(3).lower()
    month = _MONTH_NUM.get(month_name)
    if not month:
        return None
    year = start_year + _MONTH_YEAR_OFFSET.get(sheet_name, 0)
    try:
        return date(year, month, day_num)
    except ValueError:
        return None


def _detect_start_year(workbook):
    """Read Year from the Readme sheet when present; default 2025."""
    if "Readme" not in workbook.sheetnames:
        return 2025
    ws = workbook["Readme"]
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=12, values_only=True):
        vals = list(row)
        for i, cell in enumerate(vals):
            if isinstance(cell, str) and cell.strip().lower() == "year:":
                for nxt in vals[i + 1:]:
                    if isinstance(nxt, int) and 2000 <= nxt <= 2100:
                        return nxt
                    if isinstance(nxt, str) and nxt.strip().isdigit():
                        return int(nxt.strip())
    return 2025


def _map_kind(type_value):
    if not type_value:
        return None
    cleaned = str(type_value).strip().lower()
    if cleaned in _IGNORE_TYPES:
        return "__ignore__"
    return _KIND_MAP.get(cleaned)


def _parse_staff_tokens(staff_raw):
    if staff_raw in (None, ""):
        return []
    text = str(staff_raw).strip()
    if not text:
        return []
    # Pickup rows often put a return date in the Staff column.
    if DATE_LIKE.match(text) and "/" in text and len(text) < 24:
        return []
    # Drop vehicle / equipment parentheticals (complete and truncated).
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\([^)]*$", " ", text)
    text = re.sub(r"^[^(]*\)", " ", text)
    tokens = []
    for tok in re.split(r"[/,\s]+", text):
        tok = tok.strip().strip(".-+!")
        if not tok or len(tok) < 2:
            continue
        if "(" in tok or ")" in tok:
            continue
        if tok.isdigit() or tok[0].isdigit():
            continue
        if tok.lower() in JUNK_TOKENS:
            continue
        # Names are alphabetic (optionally with an initial suffix like AngusA).
        if not re.fullmatch(r"[A-Za-z][A-Za-z']*[A-Za-z0-9]*", tok):
            continue
        if VEHICLE_HINTS.fullmatch(tok):
            continue
        if DATE_LIKE.match(tok):
            continue
        # Reject datetime leftovers / time crumbs like "4pm".
        if re.match(r"^\d{4}-\d{2}-\d{2}", tok):
            continue
        if re.fullmatch(r"\d{1,2}(:\d{2})?(am|pm)", tok, re.I):
            continue
        tokens.append(tok)
    return tokens


def load_roster(path=None):
    """Parse the schedule workbook into a RosterIndex. Returns empty index if missing."""
    index = RosterIndex()
    try:
        import openpyxl
    except ImportError:
        index.load_error = (
            "openpyxl is not installed in this Python environment. "
            "Run: pip install openpyxl==3.1.5"
        )
        logger.error(index.load_error)
        return index

    roster_path = find_roster_path(path)
    if not roster_path:
        index.load_error = (
            "No Schedule*.xlsx found in the project folder. "
            f"Expected something like '{DEFAULT_ROSTER_FILENAME}'."
        )
        logger.info(index.load_error)
        return index

    try:
        workbook = openpyxl.load_workbook(roster_path, read_only=True, data_only=True)
    except Exception as exc:
        index.load_error = f"Failed to open roster workbook {roster_path}: {exc}"
        logger.error(index.load_error)
        return index

    start_year = _detect_start_year(workbook)
    for sheet_name in MONTH_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        colmap = None
        current_day = None
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            first = next((c for c in vals if c not in (None, "")), None)
            if isinstance(first, str) and DAY_HEADER_RE.match(first.strip()):
                current_day = _parse_day_header(first, sheet_name, start_year)
                colmap = None
                continue
            if vals and vals[0] == "Arrive":
                colmap = {
                    str(c).strip().lower(): i
                    for i, c in enumerate(vals)
                    if c is not None
                }
                continue
            if colmap is None:
                continue

            def get(name):
                i = colmap.get(name)
                return vals[i] if (i is not None and i < len(vals)) else None

            suburb = get("suburb")
            if not suburb:
                continue
            suburb_s = str(suburb).strip()
            if suburb_s.lower() in SKIP_SUBURBS or len(suburb_s) < 3:
                continue

            kind = _map_kind(get("type"))
            if kind == "__ignore__":
                continue

            job_match = JOBNUM_RE.search(suburb_s)
            if not job_match:
                continue
            job_number = job_match.group(1).upper()
            staff_tokens = _parse_staff_tokens(get("staff"))
            if not staff_tokens:
                continue
            index.add(job_number, current_day, kind, staff_tokens, suburb_s)

    index.source_path = roster_path
    index.loaded_at = datetime.now()
    logger.info(
        "Loaded roster from %s (%d entries, %d nicknames)",
        roster_path,
        index.entry_count,
        len(index.nicknames),
    )
    return index


def resolve_nickname(nickname, aliases, employees):
    """Map an Excel nickname to a timesheet team_member name, or None."""
    if not nickname:
        return None
    nick = str(nickname).strip()
    if not nick:
        return None

    # Explicit saved alias wins.
    if nick in aliases and aliases[nick] in employees:
        return aliases[nick]
    # Case-insensitive alias key lookup.
    nick_lower = nick.lower()
    for key, value in aliases.items():
        if key.lower() == nick_lower and value in employees:
            return value

    # Exact full-name match against timesheet employees.
    for emp in employees:
        if emp.lower() == nick_lower:
            return emp

    return None


def suggest_employee(nickname, employees):
    """Best unique auto-suggestion for the admin dropdown (not a hard mapping)."""
    exact = resolve_nickname(nickname, {}, employees)
    if exact:
        return exact
    nick_lower = str(nickname or "").strip().lower()
    if not nick_lower:
        return ""
    first_hits = [
        emp for emp in employees
        if emp.split() and emp.split()[0].lower() == nick_lower
    ]
    if len(first_hits) == 1:
        return first_hits[0]
    return ""


def build_allowed_names(nicknames, aliases, employees):
    """Resolve a set of Excel nicknames to timesheet full names."""
    allowed = set()
    for nick in nicknames:
        resolved = resolve_nickname(nick, aliases, employees)
        if resolved:
            allowed.add(resolved)
            continue
        # Ambiguous / unmapped first-name fallback: any employee whose first
        # token matches the nickname (case-insensitive). Lets "Daniel" match
        # "Daniel Hagger" before an alias is saved; ambiguous nicknames like
        # "Nick" expand to every Nicholas* — still better than no filter.
        nick_lower = str(nick).lower()
        for emp in employees:
            first = emp.split()[0].lower() if emp.split() else ""
            if first == nick_lower:
                allowed.add(emp)
    return allowed


# Module-level cache populated lazily / on admin reload.
_roster_cache = None
_aliases_cache = None
_roster_lock = RLock()
_aliases_lock = RLock()


def get_roster():
    global _roster_cache
    if _roster_cache is None:
        with _roster_lock:
            if _roster_cache is None:
                refresh_roster()
    return _roster_cache


def get_aliases():
    global _aliases_cache
    if _aliases_cache is None:
        with _aliases_lock:
            if _aliases_cache is None:
                refresh_aliases()
    return dict(_aliases_cache)


def refresh_roster(path=None):
    global _roster_cache
    loaded = load_roster(path)
    with _roster_lock:
        _roster_cache = loaded
    return loaded


def refresh_aliases(path=ALIASES_PATH):
    global _aliases_cache
    loaded = load_aliases(path)
    with _aliases_lock:
        _aliases_cache = loaded
    return loaded


def update_aliases(aliases, path=ALIASES_PATH):
    global _aliases_cache
    saved = save_aliases(aliases, path)
    with _aliases_lock:
        _aliases_cache = saved
    return saved
